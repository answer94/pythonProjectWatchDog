from datetime import datetime

from openpyxl import load_workbook
file = 'log.xlsx'


def add_path(filepath):
    wb = load_workbook(filename=file)
    ws = wb.active
    ws.insert_rows(2, 1)
    d = ws.cell(row=2, column=1, value=datetime.now())
    d = ws.cell(row=2, column=2, value="добавлен")
    d = ws.cell(row=2, column=3, value=filepath)
    wb.save(file)
    wb.close()


def del_path(filepath):
    wb = load_workbook(filename=file)
    ws = wb.active
    ws.insert_rows(2)
    d = ws.cell(row=2, column=1, value=datetime.now())
    d = ws.cell(row=2, column=2, value="удален")
    d = ws.cell(row=2, column=3, value=filepath)
    wb.save(file)
    wb.close()

def first_start():
    wb = load_workbook(filename=file)
    ws = wb.active
    ws.insert_rows(2)
    d = ws.cell(row=2, column=1, value=datetime.now())
    d = ws.cell(row=2, column=2, value="Запуск")
    wb.save(file)
    wb.close()

def end_start():
    wb = load_workbook(filename=file)
    ws = wb.active
    ws.insert_rows(2)
    d = ws.cell(row=2, column=1, value=datetime.now())
    d = ws.cell(row=2, column=2, value="Запуск")
    wb.save(file)
    wb.close()
